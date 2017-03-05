from openpyxl import load_workbook
wb = load_workbook('/home/metal-machine/Desktop/web scrapped file.xlsx')
#print wb.get_sheet_names()
#[u'Sheet1', u'Sheet2', u'Sheet3']
sheet_ranges = wb['Sheet1']

for i in range(9,20):
    print sheet_ranges.cell(row = i, column = 9).value



